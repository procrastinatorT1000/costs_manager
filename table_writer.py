import openpyxl
import os.path


def init_table(table_path = 'records.xlsx'):
   print ('Initialize table ' + table_path )

   if os.path.isfile(table_path):
      print("File exists open it")

      workbook = openpyxl.load_workbook('records.xlsx')
      sheet_list = workbook.sheetnames
      worksheet = workbook[sheet_list[0]]
   else:
      print("File not exists create it")

      workbook = openpyxl.Workbook()
      worksheet = workbook.active

   return workbook, worksheet

def write_record( sheet_item, row_lst ):

   empty_row_id = 1;
   while True:
      if (sheet_item['A'+str(empty_row_id)].value ==
         sheet_item['B'+str(empty_row_id)].value == 
         sheet_item['C'+str(empty_row_id)].value):
         break
      else:
         empty_row_id +=1
   

   print ("Saving record to records.xlsx" + str(row_lst) + "in " + str(empty_row_id) + " row")
   
   column_names = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
   column_idx = 0;
   for cell in row_lst:
      column = column_names[column_idx]
      sheet_item[column+str(empty_row_id)] = cell;
      column_idx += 1;

def deinit_table(book_item):
   print ("Deinit table")
   book_item.save('records.xlsx')

def test_write():
   book, sheet = init_table('records.xlsx')
   write_record( sheet, [1,23,'hui', 34, '77777777777777777777777', 'bye'])
   deinit_table( book )

if __name__ == '__main__':
   test_write()
